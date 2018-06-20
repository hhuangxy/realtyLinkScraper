from time import sleep
from lxml import etree
import requests
import re
import csv
import datetime
import os
import openpyxl

def genPyldArea (area):
  """Generates payload - area
  """

  baseUrl = 'http://www.realtylink.org/prop_search/AreaSelect.cfm?Branch=True'

  areaLut = {
    'burnaby'        : 11,
    'coquitlam'      : 16,
    'newwestminster' : 12,
    'portcoquitlam'  : 17,
    'portmoody'      : 15,
    'richmond'       : 13,
    'vancouvereast'  : 10,
    'vancouverwest'  :  9,
    'westvancouver'  : 19,
  }

  # Look up area value
  key = area.lower()
  if key in areaLut:
    imdp = areaLut[key]
  else:
    raise

  # Post and generate area ID list
  r = requests.post(baseUrl, data={'ERTA' : imdp}, headers={'User-Agent' : 'Mozilla/5.0'})
  match = re.search(r'AIDL=(.*?)&', r.url, re.I)
  if match is not None:
    AIDL = match.group(1)
  else:
    raise

  return {'imdp': imdp, 'AIDL' : AIDL}


def genPyldType (type):
  """Generates payload - type
  """

  typeLut = {
    'apartment' : 1,
    'townhouse' : 2,
    'house'     : 5,
  }

  # Look up type value
  key = type.lower()
  if key in typeLut:
    ptytid = typeLut[key]
  else:
    raise

  return {'PTYTID' : ptytid}


def generatePayload (area, mnage, mxage, mnbd, mnbt, type, mnprc, mxprc):
  """Generates payload
  """

  payload = {
    'SRTB'  : 'P_Price',
    'ERTA'  : 'True',
    'MNAGE' : mnage,
    'MXAGE' : mxage,
    'MNBD'  : mnbd,
    'MNBT'  : mnbt,
    'MNPRC' : mnprc,
    'MXPRC' : mxprc,
    'SCTP'  : 'RS',
    'BCD'   : 'GV',
    'RSPP'  : '5',
  }

  # Return combined dictionaries together
  return {**payload, **genPyldArea(area), **genPyldType(type)}


def generateDetails (html):
  """Generates list of details URLs
  """

  baseUrl = 'http://www.realtylink.org/prop_search/Detail.cfm?'
  rExp    = re.compile(r'MLS=([a-z0-9]+)', re.I)
  details = set()

  listRaw = html.xpath("//@href[starts-with(., 'Detail.cfm')]")
  for raw in listRaw:
    match = rExp.search(raw)
    if match is not None:
      details.add(baseUrl + match.group(0))

  # Return set as list
  return list(details)


def generateNext (html):
  """Generates next URL
  """

  baseUrl = 'http://www.realtylink.org/prop_search/'
  next    = None

  img = html.find(".//img[@src='images/property_next.gif']")
  if img is not None:
    a = img.getparent()
    next = baseUrl + a.attrib['href']

  return next


def getPage (url, payload=None):
  """Go get page
  """

  # Delay to prevent being blocked
  sleep(0.1)

  page = requests.get(url, params=payload, headers={'User-Agent' : 'Mozilla/5.0'})
  print('.', end='', flush=True)

  return page


def stripChar (line):
  """Strip unwanted characters from line
  """

  newLine =  ''.join(c if (32 <= ord(c) and ord(c) <= 255) else ' ' for c in line)
  newLine = ' '.join(newLine.strip(',').split())

  return newLine


def parseInfo (infoDesc, listInfo):
  """Parse main list of info
  """

  # Initialize dictionary
  info = {d: 'NA' for d in infoDesc}

  # Find first and mls index
  fstIdx = 0
  mlsIdx = 0
  for i, val in enumerate(listInfo):
    if (fstIdx == 0) and (val != ''):
      fstIdx = i

    if val.lower() == 'mls':
      mlsIdx = i
      break

  # Fill description
  if fstIdx == mlsIdx:
    info['description'] = 'NA'
  else:
    info['description'] = ' '.join(listInfo[fstIdx:mlsIdx])
    info['description'] = ' '.join(info['description'].split()).lstrip(': ')

  # Fill dictionary
  for key, val in zip(listInfo[mlsIdx::2], listInfo[mlsIdx+1::2]):
    key = key.lower().replace(':', '')

    if key in infoDesc:
      info[key] = val

  return info


def parsePage (html):
  """Parse page
  """

  infoDesc = [
    'mls',
    'finished floor area',
    'property type',
    'lot frontage',
    'basement',
    'lot depth',
    'bedrooms',
    'age',
    'bathrooms',
    'maintenance fee'
  ]

  info = {}

  # Get address and price
  listRaw = html.xpath('//b/text()')
  listRaw = [stripChar(l) for l in listRaw]
  info['address'] = listRaw[1]
  info['price']   = listRaw[2]

  # Get features
  listRaw = html.xpath('//li/text()')
  listRaw = [stripChar(l).capitalize() for l in listRaw]
  info['features'] = ', '.join(sorted(listRaw))

  # Get other info
  listRaw = html.xpath('//font/text()')
  listRaw = [stripChar(l) for l in listRaw]
  info = {**info, **parseInfo(infoDesc, listRaw)}

  # Clean up
  rExp = re.compile(r'(:|,|sqft\.?)', re.I)
  for key, val in info.items():
    if val.lower() == 'not available':
      val = 'NA'
    elif key == 'bathrooms':
      val = rExp.sub(' ', val).split()

      try:
        t = int(val[1])
      except:
        t = 0

      try:
        h = int(val[3])
      except:
        h = 0

      val = t + (h*0.5)
    elif key not in ['address', 'description', 'features']:
      val = rExp.sub('', val).split()
      val = ' '.join(val)

    info[key] = val

  return info


def writeCsv (fName, listDict):
  """Turn list of dictionary into CSV
  """

  # Open file
  with open(fName, 'w', newline='') as csvfile:
    cwr = csv.writer(csvfile)

    # Get header
    keys = [k.title() for k in list(listDict[0])]
    keys = sorted(keys)
    cwr.writerow(keys)

    # Build/write rows
    for d in listDict:
      row = [d[k.lower()] for k in keys]
      cwr.writerow(row)

  return 'Ok!'


def writeXl (fName, listDict):
  """Turn list of dictionary into Excel
  """

  # String to Value
  s2v = lambda s: int(s) if ('.' not in s) else float(s.replace('$',''))
  m2f = lambda l, s: l.insert(0, l.pop(l.index(s)))
  m2e = lambda l, s: l.append(l.pop(l.index(s)))

  # Open file
  wb = openpyxl.Workbook()
  ws = wb.active

  # Get and write header
  keys = [k.title() for k in list(listDict[0])]
  keys.append('Maintenance Fee/Sq Ft')
  keys.append('Price/Sq Ft')
  keys = sorted(keys)

  # Shuffle
  m2f(keys, 'Bedrooms')
  m2f(keys, 'Age')
  m2f(keys, 'Area')
  m2f(keys, 'Property Type')

  m2e(keys, 'Address')
  m2e(keys, 'Mls')
  m2e(keys, 'Basement')
  m2e(keys, 'Lot Depth')
  m2e(keys, 'Lot Frontage')
  m2e(keys, 'Description')
  m2e(keys, 'Features')

  # Append row
  ws.append(keys)

  # Get column indices
  cIdx = {k : keys.index(k) + 1 for k in keys}
  cFSF = cIdx['Maintenance Fee/Sq Ft']
  cPSF = cIdx['Price/Sq Ft']

  # Get column letters
  cLet  = {ci : openpyxl.utils.get_column_letter(cIdx[ci]) for ci in cIdx}
  clFFA = cLet['Finished Floor Area']
  clFee = cLet['Maintenance Fee']
  clPrc = cLet['Price']

  # Get format indices
  cUrl  = cIdx['Url']
  cNums = [cIdx[i] for i in ['Age', 'Bedrooms', 'Finished Floor Area']]
  cDola = [cIdx[i] for i in ['Maintenance Fee', 'Maintenance Fee/Sq Ft', 'Price', 'Price/Sq Ft']]

  # Build/write rows
  r = 2
  for d in listDict:
    # Append row
    row = [d[k.lower()] if k.lower() in d else 0 for k in keys]
    ws.append(row)

    # Format URL
    ws.cell(row=r, column=cUrl).hyperlink = ws.cell(row=r, column=cUrl).value
    ws.cell(row=r, column=cUrl).style = 'Hyperlink'

    # Calculate Maintenance Fee/Sq Ft
    ws.cell(row=r, column=cFSF).value = '=%s%s/%s%s' % (clFee, r, clFFA, r)

    # Calculate Price/Sq Ft
    ws.cell(row=r, column=cPSF).value = '=%s%s/%s%s' % (clPrc, r, clFFA, r)

    # Format numbers
    for c in cNums+cDola:
      s = ws.cell(row=r, column=c).value
      try:
        ws.cell(row=r, column=c).value = s2v(s)
      except:
        ws.cell(row=r, column=c).value = s

    # Format dollars
    for c in cDola:
      ws.cell(row=r, column=c).number_format = '$#,##0.00_);[Red]($#,##0.00)'

    # Next row
    r += 1

  wb.save(fName)
  return 'Ok!'


def traversePages (area, mnage, mxage, mnbd, mnbt, type, mnprc, mxprc):
  """Dive deep into the web
  """

  # Build payload
  payload = generatePayload(area, mnage, mxage, mnbd, mnbt, type, mnprc, mxprc)

  numBytes    = 0
  listDetails = []
  nextUrl     = 'http://www.realtylink.org/prop_search/Summary.cfm'
  log         = [('%s %s' % (area, type)).title()]

  # Get first page
  page = getPage(nextUrl, payload)
  log.append(page.url.replace('%2C', ','))

  # Parse and get the next page
  while True:
    # Calculate number of bytes
    numBytes += len(page.text)

    # Convert to HTML
    html = etree.HTML(page.text)

    # Parse HTML
    listDetails += generateDetails(html)
    nextUrl = generateNext(html)
    if nextUrl is not None:
      page = getPage(nextUrl)
    else:
      break

  # Get each page with details
  info = []
  for detUrl in listDetails:
    page = getPage(detUrl)
    numBytes += len(page.text)
    html = etree.HTML(page.text)
    info.append(parsePage(html))
    info[-1]['area'] = area.title()
    info[-1]['url']  = page.url

  # Number of KBs used
  log.append('    Number of KBs: %f' % (numBytes / 1024))

  # Check empty info
  if not info:
    log.append('    No properties were found')
  else:

  return log, info


if __name__ == '__main__':
  # Output directory
  timeStamp = datetime.datetime.now().strftime('%Y_%m_%d_%H_%M_%S')
  os.mkdir(timeStamp)

  # Area list
  areaList = ['burnaby', 'coquitlam', 'newWestminster', 'portCoquitlam', 'portMoody', 'richmond', 'vancouverEast', 'vancouverWest', 'westVancouver']

  # Type list
  typeList = ['apartment', 'townhouse', 'house']

  # Min/max age
  MNAGE = 0
  MXAGE = 35

  # Min bedroom
  MNBD = 2

  # Min bathroom
  MNBT = 0

  # Min/max price
  MNPRC = 0
  MXPRC = 600000

  log  = []
  info = []
  for AREA in areaList:
    for TYPE in typeList:
      tLog, tInfo = traversePages(AREA, MNAGE, MXAGE, MNBD, MNBT, TYPE, MNPRC, MXPRC)
      log  += tLog
      info += tInfo

  # Setup logging
  with open('%s/%s_log.txt' % (timeStamp, timeStamp), 'w') as fLog:
    fLog.write('\n'.join(log))

  # Create csv
  if info:
    #writeCsv('%s/%s.csv' % (timeStamp, timeStamp), info)
    writeXl('%s/%s.xlsx' % (timeStamp, timeStamp), info)

  print('')
  print('--- D O N E ---')
